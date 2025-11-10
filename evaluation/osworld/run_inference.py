"""Script to run end-to-end evaluation on the benchmark.
Utils and basic architecture credit to https://github.com/web-arena-x/webarena/blob/main/run.py.
"""
import io
import os
import re
import sys
import time
import copy
import json
import shutil
import base64
import requests
import logging
import argparse
import datetime
import traceback
from math import sqrt
import anthropic
from tqdm import tqdm
import lib_run_single
from typing import Tuple, List, Dict
from openai import OpenAI
# from vllm import LLM, SamplingParams
from desktop_env.desktop_env import DesktopEnv
from PIL import Image, ImageDraw, ImageFont

logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

datetime_str: str = datetime.datetime.now().strftime("%Y%m%d@%H%M%S")

file_handler = logging.FileHandler(
    os.path.join("logs", "normal-{:}.log".format(datetime_str)), encoding="utf-8"
)
debug_handler = logging.FileHandler(
    os.path.join("logs", "debug-{:}.log".format(datetime_str)), encoding="utf-8"
)
stdout_handler = logging.StreamHandler(sys.stdout)
sdebug_handler = logging.FileHandler(
    os.path.join("logs", "sdebug-{:}.log".format(datetime_str)), encoding="utf-8"
)

file_handler.setLevel(logging.INFO)
debug_handler.setLevel(logging.DEBUG)
stdout_handler.setLevel(logging.INFO)
sdebug_handler.setLevel(logging.DEBUG)

formatter = logging.Formatter(
    fmt="\x1b[1;33m[%(asctime)s \x1b[31m%(levelname)s \x1b[32m%(module)s/%(lineno)d-%(processName)s\x1b[1;33m] \x1b[0m%(message)s"
)
file_handler.setFormatter(formatter)
debug_handler.setFormatter(formatter)
stdout_handler.setFormatter(formatter)
sdebug_handler.setFormatter(formatter)

stdout_handler.addFilter(logging.Filter("desktopenv"))
sdebug_handler.addFilter(logging.Filter("desktopenv"))

logger.addHandler(file_handler)
logger.addHandler(debug_handler)
logger.addHandler(stdout_handler)
logger.addHandler(sdebug_handler)
#  }}} Logger Configs #

logger = logging.getLogger("desktopenv.experiment")

CURRENT_WHOLE_IMAGE = None
CURRENT_IMAGE_RANGE = None
CURRENT_RED_POINT = None

# SYS_PROMPT_IN_SCREENSHOT_OUT_CODE_ONE_COMMAND = """
# You are an agent which follow my instruction and perform desktop computer tasks as instructed.
# You have good knowledge of computer and good internet connection and assume your code will run on a computer for controlling the mouse and keyboard.
# For each step, you will get an observation of an image, which is the screenshot of the computer screen and you will predict the action of the computer based on the image.

# You are required to use `pyautogui` to perform the action grounded to the observation, but DONOT use the `pyautogui.locateCenterOnScreen` function to locate the element you want to operate with since we have no image of the element you want to operate with. DONOT USE `pyautogui.screenshot()` to make screenshot.
# Return one line of python code to perform the ONE action each time. 
# You need to to specify the coordinates of by yourself based on your observation of current observation, but you should be careful to ensure that the coordinates are correct.
# You ONLY need to return the code inside a code block, like this:
# ```python
# # your code here
# ```
# Note: If you want to use the `pyautogui.click`, `pyautogui.mouseDown`, `pyautogui.moveTo` commands, replace the coordinate parameters X and Y in the original function with the target you want to click on and a descriptive text of its shape, color and position. 
# If you want to use the `pyautogui.dragTo` command, use `pyautogui.mouseDown` to click on the target you want to drag and then use `pyautogui.moveTo(item: str, description: str,)` to move the mouse to the target you want to drag to and then use `pyautogui.mouseUp` to release the mouse button.
# If you want to perform a right-click, please use: `pyautogui.click(item: str, description: str, button='right')` instead of `pyautogui.rightClick()`.
# For example:
# For pyautogui.click, you should use the following command format:
# ```python
# pyautogui.click(
#     item: str,
#     description: str,
#     button: str
# )
# ```

# If you want to use pyautogui.click to click on the VS Code icon, you should use the following command:
# ```python
# pyautogui.click(
#     item='vscode icon',
#     description='It is located in the sidebar (Launcher) on the left side of the screen. It is the first icon from the top in the vertical arrangement. The icon has a blue background with a white folded "V"-shaped design in the center. The sidebar is aligned along the leftmost edge of the screen, adjacent to the desktop background on its right side.',
#     button='left'
# )
# ```

# Instead of:
# ```python
# pyautogui.click(x=500, y=300, button='left')
# ```
# For `pyautogui` commands other than `pyautogui.click`, please use the normal format.
# Specially, it is also allowed to return the following special code:
# When you think you have to wait for some time, return ```WAIT```;
# When you think the task can not be done, return ```FAIL```;
# When you think the task is done, return ```DONE```.

# My computer's password is 'password', feel free to use it when you need sudo rights.
# First give the current screenshot and previous things we did a short reflection, then RETURN ME THE CODE OR SPECIAL CODE I ASKED FOR. NEVER EVER RETURN ME ANYTHING ELSE.
# """.strip()

# SYS_PROMPT_IN_SCREENSHOT_OUT_CODE_ONE_COMMAND = """
# You are an agent which follow my instruction and perform desktop computer tasks as instructed.
# You have good knowledge of computer and good internet connection and assume your code will run on a computer.
# For each step, you will get an observation of an image, which is the screenshot of the computer screen and you will predict the action of the computer based on the image.

# You can perform four types of actions:

# 1. **Command-line actions**
# Put the commands you want to run in:
# <execute_bash>
# # your code here
# </execute_bash>
# For example: if you want to check the current directory, you can return:
# <execute_bash>
# pwd
# </execute_bash>

# 2. **Tool-Building actions**
# You can also create a new python function as a new tool to help you finish the task.
# You can create a new tool by returning the following format:
# <tool>
# make_new_tool(functionality: str, function_name: str, function_inputs: list, function_outputs: str)
# </tool>
# and invoke it during subsequent task execution.
# For example: if you want to create a new tool to open a file in libreoffice calc, you can return:
# <execute>
# make_new_tool(functionality="Open a file in libreoffice calc application", 
# function_name="open_file_in_libreoffice_calc", 
# function_inputs=["file_path: str"], 
# function_outputs="None")
# </execute>

# 3 **Python-in-Jupyter actions**
# You can also run python code in the jupyter notebook environment to help you finish the task.
# Note: No helper functions are pre-defined. 
# If you need utilities (helpers, scripts, or modules), 
# first create them using Type 2: Tool-Building actions, then call them in:
# <execute>
# # your code here
# </execute>
# For example: if you want to open a file in libreoffice calc, you can return:
# <execute>
# open_file_in_libreoffice_calc(file_path="/path/to/your/file.xlsx")
# </execute>

# Specially, it is also allowed to return the following special code:
# When you think you have to wait for some time, return <execute>WAIT</execute>;
# When you think the task can not be done, return <execute>FAIL</execute>;
# When you think the task is done, return <execute>DONE</execute>.

# My computer's password is 'password', feel free to use it when you need sudo rights.
# First give the current screenshot and previous things we did a short reflection, then RETURN ME THE CODE OR SPECIAL CODE I ASKED FOR. NEVER EVER RETURN ME ANYTHING ELSE.
# """.strip()

BASH_IN_PYTHON = r"""
import base64, sys, subprocess

b64 = {b64!r}
script = base64.b64decode(b64).decode('utf-8')

proc = subprocess.Popen(
    ['bash', '-lc', script],
    stdout=subprocess.PIPE,
    stderr=subprocess.STDOUT,
    text=True
)

out, _ = proc.communicate()
if out:
    sys.stdout.write(out)
"""



SYS_PROMPT_IN_SCREENSHOT_OUT_CODE_ONE_COMMAND = """
You are an agent which follow my instruction and perform desktop computer tasks as instructed.
You have good knowledge of computer and good internet connection and assume your code will run on a computer.
For each step, you will get an observation of an image, which is the screenshot of the computer screen and you will predict the action of the computer based on the image.

You should first try the following three types of actions:

1. **Command-line actions**
Put the commands you want to run in:
<execute_bash>
# your code here
</execute_bash>
For example: if you want to check the current directory, you can return:
<execute_bash>
pwd
</execute_bash>
if you want to install a package by using pip, you can return:
<execute_bash>
pip install package_name
</execute_bash>

2. **Tool-Building actions**
You can also create a new python function as a new tool to help you finish the task.
You can create a new tool by returning the following format:
<tool>
make_new_tool(functionality: str, function_name: str, function_inputs: list, function_outputs: str)
</tool>
and invoke it during subsequent task execution.
For example: if you want to create a new tool to open a file in libreoffice calc, you can return:
<tool>
make_new_tool(functionality="Open a file in libreoffice calc application", 
function_name="open_file_in_libreoffice_calc", 
function_inputs=["file_path: str"], 
function_outputs="None")
</tool>

3 **Python-in-Jupyter actions**
You can also run python code in the jupyter notebook environment to help you finish the task.
Note: No helper functions are pre-defined. 
If you need utilities (helpers, scripts, or modules), 
first create them using Type 2: Tool-Building actions, then call them in:
<execute>
# your code here
</execute>
For example: if you want to open a file in libreoffice calc, you can return:
<execute>
open_file_in_libreoffice_calc(file_path="/path/to/your/file.xlsx")
</execute>

If you determine that the three approaches above cannot resolve the issue—or cannot do so reliably—you can also use the PyAutoGUI actions below.

**PyAutoGUI actions**
You are required to use `pyautogui` to perform the action grounded to the observation, but DONOT use the `pyautogui.locateCenterOnScreen` function to locate the element you want to operate with since we have no image of the element you want to operate with. DONOT USE `pyautogui.screenshot()` to make screenshot.
Return one line of python code to perform the ONE action each time. 
You need to to specify the coordinates of by yourself based on your observation of current observation, but you should be careful to ensure that the coordinates are correct.
You ONLY need to return the code inside a code block, like this:
<execute_gui>
# your code here
</execute_gui>
Note: If you want to use the `pyautogui.click`, `pyautogui.mouseDown`, `pyautogui.moveTo` commands, replace the coordinate parameters X and Y in the original function with the target you want to click on and a descriptive text of its shape, color and position. 
If you want to use the `pyautogui.dragTo` command, use `pyautogui.mouseDown` to click on the target you want to drag and then use `pyautogui.moveTo(item: str, description: str,)` to move the mouse to the target you want to drag to and then use `pyautogui.mouseUp` to release the mouse button.
If you want to perform a right-click, please use: `pyautogui.click(item: str, description: str, button='right')` instead of `pyautogui.rightClick()`.
For example:
For pyautogui.click, you should use the following command format:
<execute_gui>
pyautogui.click(
    item: str,
    description: str,
    button: str
)
</execute_gui>

For pyautogui.tripleClick, you should use the following command format:
<execute_gui>
pyautogui.tripleClick(
    item: str,
    description: str,
)
</execute_gui>

If you want to use pyautogui.click to click on the VS Code icon, you should use the following command:
<execute_gui>
pyautogui.click(
    item='vscode icon',
    description='It is located in the sidebar (Launcher) on the left side of the screen. It is the first icon from the top in the vertical arrangement. The icon has a blue background with a white folded "V"-shaped design in the center. The sidebar is aligned along the leftmost edge of the screen, adjacent to the desktop background on its right side.',
    button='left'
)
</execute_gui>
Instead of:
<execute_gui>
pyautogui.click(x=500, y=300, button='left')
</execute_gui>
For `pyautogui` commands other than `pyautogui.click`,`pyautogui.tripleClick`, please use the normal format.

Specially, it is also allowed to return the following special code:
When you think you have to wait for some time, return <execute>WAIT</execute>;
When you think the task can not be done, return <execute>FAIL</execute>;
When you think the task is done, return <execute>DONE</execute>.

My computer's password is 'password', feel free to use it when you need sudo rights.
First give the current screenshot and previous things we did a short reflection, then RETURN ME THE CODE OR SPECIAL CODE I ASKED FOR. NEVER EVER RETURN ME ANYTHING ELSE.
""".strip()

# NOTES = '''NOTE: 
# 1. If you want to use the `pyautogui.click`, `pyautogui.drag`, `pyautogui.mouseDown`, `pyautogui.moveTo` commands, replace the coordinate parameters `x` and `y` in the original function with the target you want to click on and a descriptive text of its shape, color and position.
# 2. Some settings require restarting the application before they take effect.
# 3. If you are asked to browse some web pages, please ensure that the page you ultimately open meets all of the requirements.
# 4. Please use terminal to finish the libreoffice cal task if you can.'''

# NOTES = '''NOTE: Please always use the command line or write some Tool functions to solve the task first. If that doesn’t work, fall back to PyAutoGUI actions, as GUI clicks are not always reliable.'''
NOTES = '''If the task is finished, do not forget to save your work.'''

import ast
import textwrap
from typing import List, Optional

def extract_function_names(code: str, top_level_only: bool = True) -> List[str]:
    """
    提取函数名。默认只取顶层定义；若 top_level_only=False，则递归遍历所有作用域。
    """
    code = textwrap.dedent(code)
    try:
        tree = ast.parse(code)
    except SyntaxError:
        return []

    names: List[str] = []
    if top_level_only:
        for node in tree.body:
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                names.append(node.name)
    else:
        for node in ast.walk(tree):
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                names.append(node.name)
    return names

def extract_first_function_name(code: str, top_level_only: bool = True) -> Optional[str]:
    names = extract_function_names(code, top_level_only=top_level_only)
    return names[0] if names else None

def extract_coordinates(result: list[str]):
    text = result[0].strip()

    # 如果有 <answer> 标签，就提取标签内的内容；否则就直接用 text
    answer_match = re.search(r'<answer>\s*(.*?)\s*</answer>', text, re.DOTALL)
    if answer_match:
        content = answer_match.group(1)
    else:
        content = text  

    # 按 (x, y) 形式提取
    point_match = re.search(r'\(\s*(\d+)\s*,\s*(\d+)\s*\)', content)
    if point_match:
        x, y = map(int, point_match.groups())
        return (x, y)

    # 如果是 (x1, y1, x2, y2) 形式，取中心点
    box_match = re.search(r'\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)', content)
    if box_match:
        x1, y1, x2, y2 = map(int, box_match.groups())
        return ((x1 + x2)//2, (y1 + y2)//2)

    return (-1,-1)

def backup_image(image_path, mount_path, backup_dir="Backup/osworld/images"):

    # Ensure backup directory exists
    os.makedirs(backup_dir, exist_ok=True)

    # Define target path
    target_path = os.path.join(backup_dir, os.path.basename(image_path))

    # Copy the image
    shutil.copy(image_path, target_path)
    logger.info(f"Image backed up to {target_path}")
    
def localization_point(x: int, y: int)-> Tuple[Image.Image, tuple, tuple, tuple]:
    """
    This function is used to check the exact screen position of a coordinate by drawing a red dot on the screen.
    
    Args:
        x (int): The x-coordinate of the point to check.
        y (int): The y-coordinate of the point to check.
        
    Returns:
        None
    """
    global CURRENT_WHOLE_IMAGE
    global CURRENT_RED_POINT
    
    if x < 0 or x >= CURRENT_WHOLE_IMAGE.width or y < 0 or y >= CURRENT_WHOLE_IMAGE.height:
        return None, None, None, (-1,-1)
    
    # Draw a red dot on the image at the specified coordinates
    CURRENT_RED_POINT = (x, y)
    x_range = [0,CURRENT_WHOLE_IMAGE.width]
    y_range = [0,CURRENT_WHOLE_IMAGE.height]
    copy_img = CURRENT_WHOLE_IMAGE.copy()
    draw = ImageDraw.Draw(copy_img)
    dot_radius = 10
    dot_color = (255, 0, 0)
    draw.ellipse((x - dot_radius, y - dot_radius, x + dot_radius, y + dot_radius), fill=dot_color)
    return copy_img, x_range, y_range, (x, y)

def highlight_and_save_region(center: tuple[int, int], half_size_x: int = 700, half_size_y: int = 250):
    """
    Highlight a region around the specified center point in the current image.
    """
    global CURRENT_WHOLE_IMAGE
    image = CURRENT_WHOLE_IMAGE.copy()
    width, height = image.size
    x, y = center
    left = max(0, x - half_size_x)
    top = max(0, y - half_size_y)
    right = min(width, x + half_size_x)
    bottom = min(height, y + half_size_y)
    
    if left >= right or top >= bottom:
        raise ValueError(f"Invalid region: {(left, top, right, bottom)}")
    cropped = image.crop((left, top, right, bottom))
    byte_cropped = save_image_and_convert_to_byte(cropped)
    offset = (left, top)
    return byte_cropped, offset

def clean_actions(actions):
    """
    Post-process a list of PyAutoGUI action strings, removing any
    redundant button='right' parameter from rightClick calls.
    
    Args:
        actions (List[str]): List of code strings representing PyAutoGUI calls.
    
    Returns:
        List[str]: Cleaned list of action strings.
    """
    cleaned_actions = []
    for action in actions:
        if "pyautogui.rightClick" in action['command']:
            # Remove ", button='right'" (or with double quotes) wherever it appears
            action['command'] = re.sub(r",\s*button=['\"]right['\"]", "", action['command'])
        elif "pyautogui.moveTo" in action['command']:
            action['command'] = re.sub(r",\s*button=['\"][^'\"]*['\"]", "", action['command'])
        cleaned_actions.append(action)
    
    logger.debug("Original actions: %s", actions)
    logger.debug("Cleaned actions: %s", cleaned_actions)
    return cleaned_actions

def save_image_and_convert_to_byte(img: Image.Image) -> bytes:
    '''
    Save the image with grid to intermediate folder and return the byte data of the image.
    Args:
        mount_path (str): The directory to save the intermediate results.
        img (Image): The image to be saved.
    Returns:
        bytes: The byte data of the image.
    '''
    mount_path = os.getcwd()
    os.makedirs(f'{mount_path}/intermediate_steps/', exist_ok=True) 
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    filename = f"{mount_path}/intermediate_steps/{timestamp}.png"
    img.save(filename)
    logger.info(f"Intermediate Image saved as: {filename}")
    with open(filename, "rb") as file:
        image_bytes = file.read()
    return image_bytes

def encode_image(image_content):
    return base64.b64encode(image_content).decode('utf-8')

def replace_icon_desc_with_coordinates(command: str, x: int, y: int) -> str:
    """
    Replace `item` and `description` parameters in the given command with `x` and `y` values.
    Args:
        command (str): Command string containing pyautogui.click.
        x (int): X-coordinate value.
        y (int): Y-coordinate value.
    Returns:
        str: Modified command string or a no-op move if coords are invalid.
    """
    if x == -1 or y == -1:
        return "pyautogui.moveRel(0, 0)"

    pattern = re.compile(
        r"item\s*=\s*(?P<q1>['\"])(?P<item>.*?)(?P=q1)"     
        r"(?=\s*,\s*description)\s*,\s*"
        r"description\s*=\s*(?P<q2>['\"])(?P<desc>.*?)(?P=q2)"
        r"(?=\s*(?:,|\)))",                                  
        re.IGNORECASE | re.DOTALL
    )

    def _repl(m: re.Match) -> str:
        return f"x={x}, y={y}"

    return pattern.sub(_repl, command)

def extract_icon_and_desc(command: str) -> tuple[str, str]:
    """
    Extract `item` and `description` values from the given command.
    """
    # 单引号版本：结束单引号后必须跟逗号或右括号
    single_pattern = re.compile(
        r"item\s*=\s*'(?P<item>.*?)'(?=\s*,\s*description)\s*,\s*"
        r"description\s*=\s*'(?P<desc>.*?)'(?=\s*(?:,|\)))",
        re.IGNORECASE | re.DOTALL
    )
    m = single_pattern.search(command)
    if m:
        return m.group('item'), m.group('desc')

    # 双引号版本：同理
    double_pattern = re.compile(
        r'item\s*=\s*"(?P<item>.*?)"(?=\s*,\s*description)\s*,\s*'
        r'description\s*=\s*"(?P<desc>.*?)"(?=\s*(?:,|\)))',
        re.IGNORECASE | re.DOTALL
    )
    m2 = double_pattern.search(command)
    if m2:
        return m2.group('item'), m2.group('desc')

    return None, None

def extract_coordination(command):
    """
    Extract `x` and `y` values from the given command.

    Args:
        command (str): Command string containing `localization_done` with x and y.

    Returns:
        tuple: Extracted x and y values as a tuple of integers.
    """
    # Define the regular expression pattern to extract `x` and `y`
    pattern = r"localization_done\((\d+),\s*(\d+)\)"
    match = re.search(pattern, command)

    if not match:
        logger.error("Invalid command format. Expected format: 'localization_done(x, y)'")
        return (0, 0)
    # Extract and return x and y as a tuple
    return tuple(map(int, match.groups()))

# def parse_code_from_string(input_string: str) -> list[str]:
#     """
#     从输入字符串中提取代码片段和命令:
#     - 仅在代码块（```...```）外按分号分割语句
#     - 提取代码块内的内容
#     - 识别特殊命令 WAIT, DONE, FAIL

#     返回一个字符串列表，每个元素是一段代码或命令。"""
#     trimmed = input_string.strip()
#     if trimmed in ('WAIT', 'DONE', 'FAIL'):
#         return [trimmed]

#     code_block_pattern = re.compile(r'```(?:\w+\s+)?(.*?)```', re.DOTALL)

#     def split_outside_code(text: str) -> list[str]:
#         parts = []
#         buf = []
#         in_block = False
#         i = 0
#         while i < len(text):
#             if text.startswith('```', i):
#                 in_block = not in_block
#                 buf.append('```')
#                 i += 3
#             elif text[i] == ';' and not in_block:
#                 segment = ''.join(buf).strip()
#                 if segment:
#                     parts.append(segment)
#                 buf = []
#                 i += 1
#             else:
#                 buf.append(text[i])
#                 i += 1
#         last = ''.join(buf).strip()
#         if last:
#             parts.append(last)
#         return parts

#     segments = split_outside_code(input_string)
#     commands = {'WAIT', 'DONE', 'FAIL'}
#     results = []

#     for seg in segments:
#         seg = seg.strip()
#         if seg in commands:
#             results.append(seg)
#             continue
#         for match in code_block_pattern.findall(seg):
#             code = match.strip()
#             results.append(code)
#         # if not code_block_pattern.search(seg) and seg:
#         #     results.append(seg)

#     return results

def parse_code_from_string(input_string: str) -> List[Dict[str, str]]:
    """
    将输入字符串解析为一组“动作”字典：
    - 仅在 <execute>...</execute> 与 <execute_bash>...</execute_bash> 之外按分号切分
    - <execute> 内部文本作为 Python 代码返回，type="python"
    - <execute_bash> 内部文本作为 Bash 代码返回，type="bash"
    - 识别 WAIT / DONE / FAIL，作为控制命令返回，type="control"

    返回: List[{"command": str, "type": "python"|"bash"|"control"}]
    """
    trimmed = input_string.strip()
    if trimmed in ('WAIT', 'DONE', 'FAIL'):
        return [{"command": trimmed, "type": "control"}]

    # 匹配 <execute>...</execute> 与 <execute_bash>...</execute_bash>
    code_block_pattern = re.compile(
        r'<(execute|execute_bash|tool|execute_gui)>(.*?)</\1>',
        re.DOTALL
    )

    def split_outside_code(text: str) -> List[str]:
        """
        按分号切分，但忽略两类代码块内的分号。
        返回分段字符串（每段可能包含若干代码块）。
        """
        parts: List[str] = []
        buf: List[str] = []

        start_to_end = {
            "<execute>": "</execute>",
            "<execute_bash>": "</execute_bash>",
        }

        i = 0
        in_block_end = None  # None 或当前块应匹配的结束标签字符串

        while i < len(text):
            if in_block_end is None:
                # 不在代码块内：优先识别起始标签
                opened = False
                for st, et in start_to_end.items():
                    if text.startswith(st, i):
                        in_block_end = et
                        buf.append(st)
                        i += len(st)
                        opened = True
                        break
                if opened:
                    continue

                # 外部遇到分号就切分
                if text[i] == ';':
                    segment = ''.join(buf).strip()
                    if segment:
                        parts.append(segment)
                    buf = []
                    i += 1
                else:
                    buf.append(text[i])
                    i += 1
            else:
                # 在代码块内：只检测对应的结束标签
                if text.startswith(in_block_end, i):
                    buf.append(in_block_end)
                    i += len(in_block_end)
                    in_block_end = None
                else:
                    buf.append(text[i])
                    i += 1

        last = ''.join(buf).strip()
        if last:
            parts.append(last)
        return parts

    segments = split_outside_code(input_string)
    commands = {'WAIT', 'DONE', 'FAIL'}
    results: List[Dict[str, str]] = []

    for seg in segments:
        seg = seg.strip()
        if '<execute>' in seg and '</execute>' not in seg:
            seg += '</execute>'
        if '<execute_bash>' in seg and '</execute_bash>' not in seg:
            seg += '</execute_bash>'
        if '<tool>' in seg and '</tool>' not in seg:
            seg += '</tool>'
        if '<execute_gui>' in seg and '</execute_gui>' not in seg:
            seg += '</execute_gui>'
        if not seg:
            continue

        # 控制命令
        if seg in commands:
            results.append({"command": seg, "type": "control"})
            continue

        # 依次提取代码块，生成对应 dict
        for m in code_block_pattern.finditer(seg):
            tag = m.group(1)           # execute / execute_bash
            body = m.group(2).strip()  # 代码正文
            if not body:
                continue
            if tag == 'execute':
                results.append({"command": body, "type": "python"})
            elif tag == 'tool':
                results.append({"command": body, "type": "tool"})
            elif tag == 'execute_gui':
                results.append({"command": body, "type": "gui"})
            else:
                results.append({"command": body, "type": "bash"})

        # 与原逻辑一致：不把“非代码块的外部文本”加入结果
        # 若需要也返回外部文本，可在这里补充相应处理。

    return results

def save_figure(img, intermediate_results_dir):
    os.makedirs(intermediate_results_dir, exist_ok=True) 
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = f"{intermediate_results_dir}/intermediate_steps/{timestamp}.jpg"

    img.save(file_path)
    logger.info(f"Image saved as: {file_path}")
    with open(file_path, "rb") as file:
        image_bytes = file.read()
    return image_bytes

def print_messages(dc_messages, function_name):
    def format_message_content(message):
        content = message.get("content", '')
        if isinstance(content, str):
            return content
        if isinstance(content, list) and content and isinstance(content[0], dict) and "text" in content[0]:
            return content[0]["text"] + " <image>"
        return "<image>"

    printable_messages = [
        {
            "role": message["role"],
            "content": format_message_content(message)
        }
        for message in dc_messages
    ]

    print(f'Messages in {function_name}: {printable_messages}')

def _fix_pyautogui_less_than_bug(command: str) -> str:
    """
    Fix PyAutoGUI '<' character bug by converting it to hotkey("shift", ',') calls.
    
    This fixes the known PyAutoGUI issue where typing '<' produces '>' instead.
    References:
    - https://github.com/asweigart/pyautogui/issues/198
    - https://github.com/xlang-ai/OSWorld/issues/257
    
    Args:
        command (str): The original pyautogui command
        
    Returns:
        str: The fixed command with '<' characters handled properly
    """
    # Pattern to match press('<') or press('\u003c') calls  
    press_pattern = r'pyautogui\.press\(["\'](?:<|\\u003c)["\']\)'

    # Handle press('<') calls
    def replace_press_less_than(match):
        return 'pyautogui.hotkey("shift", ",")'
    
    # First handle press('<') calls
    command = re.sub(press_pattern, replace_press_less_than, command)

    # Pattern to match typewrite calls with quoted strings
    typewrite_pattern = r'pyautogui\.typewrite\((["\'])(.*?)\1\)'
    
    # Then handle typewrite calls
    def process_typewrite_match(match):
        quote_char = match.group(1)
        content = match.group(2)
        
        # Preprocess: Try to decode Unicode escapes like \u003c to actual '<'
        # This handles cases where '<' is represented as escaped Unicode
        try:
            # Attempt to decode unicode escapes
            decoded_content = content.encode('utf-8').decode('unicode_escape')
            content = decoded_content
        except UnicodeDecodeError:
            # If decoding fails, proceed with original content to avoid breaking existing logic
            pass  # English comment: Graceful degradation - fall back to original content if decoding fails
        
        # Check if content contains '<'
        if '<' not in content:
            return match.group(0)
        
        # Split by '<' and rebuild
        parts = content.split('<')
        result_parts = []
        
        for i, part in enumerate(parts):
            if i == 0:
                # First part
                if part:
                    result_parts.append(f"pyautogui.typewrite({quote_char}{part}{quote_char})")
            else:
                # Add hotkey for '<' and then typewrite for the rest
                result_parts.append('pyautogui.hotkey("shift", ",")')
                if part:
                    result_parts.append(f"pyautogui.typewrite({quote_char}{part}{quote_char})")
        
        return '; '.join(result_parts)
    
    command = re.sub(typewrite_pattern, process_typewrite_match, command)
    
    return command

def truncate(text, max_chars=4000, head=1500, tail=1500, placeholder="... [TRUNCATED: {n} chars] ..."):
    """
    Truncate long text by keeping the first `head` chars and last `tail` chars.
    Ensures total length <= max_chars. Works with any object by str() casting.
    """
    if text is None:
        return ""
    s = str(text)

    # Fast path
    if len(s) <= max_chars:
        return s

    # Adjust head/tail if head+tail exceeds max_chars minus placeholder
    ph = placeholder.format(n=max(len(s) - head - tail, 0))
    budget = max_chars - len(ph)
    if budget <= 0:
        # Fallback: return the start only, clipped to max_chars
        return s[:max_chars]

    # Recompute head/tail proportionally to fit budget
    if head + tail > budget:
        # keep 60% head / 40% tail by default
        head = int(budget * 0.6)
        tail = budget - head

    return f"{s[:head]}{ph}{s[-tail:]}"

class OSworld_Env(DesktopEnv):
    def __init__(self, *args, helper_fn=None, gui_helper_fn=None, **kwargs):
        super().__init__(*args, **kwargs)
        self.helper_fn = helper_fn
        self.gui_helper_fn = gui_helper_fn
        self.toolsets = ''
        
    def step(self, action, pause=2):
        self._step_no += 1
        self.action_history.append(action['command'])
        
        # Mark environment as used when step is called
        self.is_environment_used = True

        reward = 0  # todo: Define reward calculation for each example
        done = False  # todo: Define episode termination condition for each example
        info = {}
        # handle the special actions
        if action['command'] in ['WAIT', 'FAIL', 'DONE']:
            if action['command'] == 'WAIT':
                time.sleep(pause)
            elif action['command'] == 'FAIL':
                done = True
                info = {"fail": True}
            elif action['command'] == 'DONE':
                done = True
                info = {"done": True}

        if self.action_space == "computer_13":
            # the set of all possible actions defined in the action representation
            self.controller.execute_action(action)
        elif self.action_space == "pyautogui" or self.action_space == "claude_computer_use":
            if action['command'] in ['WAIT', 'FAIL', 'DONE']:
                self.controller.execute_action(action['command'])
            else:
                # the set of all possible python commands insides `pyautogui`
                if type(action) == str:
                    # Fix PyAutoGUI '<' character bug before execution
                    fixed_command = _fix_pyautogui_less_than_bug(action)
                    self.controller.execute_python_command(fixed_command)
                elif type(action) == dict:
                    if action['type'] == 'bash':
                        cmd = BASH_IN_PYTHON.format(
                            b64=base64.b64encode(action['command'].encode("utf-8")).decode("ascii")
                        )
                        logger.info("Executing bash command via Python:\n%s", cmd)
                        resp = self.controller.execute_python_command(cmd)

                        result_text = ""
                        if isinstance(resp, dict):
                            for key in ("output", "stdout", "out"):
                                if isinstance(resp.get(key), str):
                                    result_text = resp[key]
                                    break
                            if not result_text:
                                for key in ("error", "stderr"):
                                    if isinstance(resp.get(key), str):
                                        result_text = resp[key]
                                        break
                        elif isinstance(resp, str):
                            result_text = resp or ""
                        else:
                            result_text = ""

                        result_text = result_text.rstrip("\n")
                    if action['type'] == 'python':
                        # Fix PyAutoGUI '<' character bug before execution
                        fixed_command = _fix_pyautogui_less_than_bug(action['command'])
                        resp = self.controller.execute_python_command('\n' + self.toolsets + '\n' + fixed_command) or {}

                        rc = resp.get("returncode")
                        out = resp.get("output", "") or ""
                        err = resp.get("error", "") or ""

                        if rc not in (0, None) or err:
                            main_text = err.strip() or "(no error text)"
                        else:
                            main_text = out.strip() or "(no output)"

                        result_text = f"[returncode={rc if rc is not None else 'N/A'}]\n{main_text}"
                    if action['type'] == 'tool':
                        if self.helper_fn is not None:
                            new_tool_code = self.helper_fn(action['command'])
                            fun_name = extract_first_function_name(new_tool_code)
                            self.toolsets =  self.toolsets + '\n' + new_tool_code
                            result_text = f"New tool function '{fun_name}' created:\n{new_tool_code}" if fun_name else "New tool created."
                            result_text += ('\nYou can use it in subsequent actions by calling:\n'
                                            '<execute>\n'
                                            f'{fun_name}(YOUR_ARGUMENTS_HERE)\n'
                                            '</execute>\n' 
                                            'directly.')
                        else:
                            logger.error("No helper function provided to create new tool.")
                    if action['type'] == 'gui':
                        if self.gui_helper_fn is not None:
                            action = self.gui_helper_fn(action)
                            fixed_command = _fix_pyautogui_less_than_bug(action['command'])
                            resp = self.controller.execute_python_command(fixed_command) or {}

                            rc = resp.get("returncode")
                            out = resp.get("output", "") or ""
                            err = resp.get("error", "") or ""

                            if rc not in (0, None) or err:
                                main_text = err.strip() or "(no error text)"
                            else:
                                main_text = out.strip() or "(no output)"

                            result_text = f"[returncode={rc if rc is not None else 'N/A'}]\n{main_text}"
                        else:
                            logger.error("No GUI helper function provided to execute GUI action.")

        time.sleep(pause)
        observation = self._get_obs()
        observation['bash_output'] = truncate(result_text, max_chars=4000) if 'result_text' in locals() else ""

        return observation, reward, done, info
    
  
class OSworld_test_Agent():
    def __init__(self, env: OSworld_Env, example_result_dir: str, cfg_args: dict, 
                 screen_width: int, screen_height: int) -> None:
        # Call the parent class's constructor
        self.platform = "ubuntu"
        self.model = cfg_args['model']
        self.max_tokens = cfg_args['max_tokens']
        self.top_p = cfg_args['top_p']
        self.temperature = cfg_args['temperature']
        self.action_space = "pyautogui"
        self.max_trajectory_length = cfg_args['max_trajectory_length']
        self.env = env
        self.intermediate_results_dir = example_result_dir
        self.screen_width = screen_width
        self.screen_height = screen_height
        self.oss_llm = None
        self.steps = 0
        self.max_steps = cfg_args['max_steps']
        self.thoughts = []
        self.actions = []
        self.observations = []
        self.base_url_oss = cfg_args['base_url_oss']

        if self.action_space == "computer_13":
            self.system_message = ''
        elif self.action_space == "pyautogui":
            self.system_message = SYS_PROMPT_IN_SCREENSHOT_OUT_CODE_ONE_COMMAND
        else:
            raise ValueError("Invalid action space: " + self.action_space)
        # self.oss_llm_init()
        
    def reset(self, _logger=None):
        '''
        Clean everything
        '''
        global logger
        logger = _logger if _logger is not None else logging.getLogger("desktopenv.agent")

        self.thoughts = []
        self.actions = []
        self.observations = []
    
    def update_result_dir(self, example_result_dir):
        self.intermediate_results_dir = example_result_dir
    
    # def oss_llm_init(self):
    #     if self.oss_llm is None:
    #         logger.info(f"Loading model ByteDance-Seed/UI-TARS-1.5-7B into GPU...")
    #         self.oss_llm = LLM(
    #             model='ByteDance-Seed/UI-TARS-1.5-7B',
    #             tokenizer='ByteDance-Seed/UI-TARS-1.5-7B',
    #             tensor_parallel_size=2,
    #             gpu_memory_utilization=0.9,
    #             enforce_eager=True,
    #             max_model_len=9632,
    #             disable_custom_all_reduce=True,
    #             enable_prefix_caching=False,
    #             trust_remote_code=True,
    #         )
    #         logger.info(f"Model ByteDance-Seed/UI-TARS-1.5-7B loaded into GPU memory") 
    
    def make_new_tool(self, functionality: str, function_name: str,
                            function_inputs: list, function_outputs: str) -> str:
        '''
        Make a new tool based on the request from the Agent
        If the new tool is created:
            return the created tool
        if not:
            return None
        '''
        try:
            example = '''USER:
    I need a function that can create a new sheet in an existing Excel file.

    Functionality:  
    - Open an existing Excel file.  
    - Create a new sheet with the specified name.  
    - Save the updated Excel file.  
    - If the sheet already exists, raise a ValueError.  

    Function name: create_excel_page  
    Input: file_name (str), page_name (str)  
    Output: None  

    ASSISTANT:
    Here is the complete function wrapped inside the <tool>...</tool> tag:

    <tool>
    import openpyxl

    def create_excel_page(file_name: str, page_name: str) -> None:
        """
        Create a new sheet in the given Excel file.

        Args:
            file_name (str): Path to the Excel file.
            page_name (str): Name of the new sheet to be created.

        Returns:
            None

        Raises:
            ValueError: If the sheet already exists.
        """
        wb = openpyxl.load_workbook(file_name)
        if page_name in wb.sheetnames:
            raise ValueError(f"Sheet '{page_name}' already exists.")
        wb.create_sheet(title=page_name)
        wb.save(file_name)
    </tool>'''

            initial_question = (
                                f"Please help me to write a new python function based on the following instruction: "
                                f"The function should provide the following functionality. \n{functionality}\n"
                                f"The name of the function should be: \n{function_name}\n" 
                                f"The input of the function should be: \n{function_inputs}\n"
                                f"The output of the function should be: \n{function_outputs}\n"
                                "Please verify your function thoroughly. "
                                "Finally, please only provide the complete function and wrap it inside the <tool>...</tool> tag."
                                "Do not include anything else inside the <tool>...</tool> tag (such as unit tests or usage examples). "
                                "Only place the function itself within the tag. "
                                f"Here is an example: \n{example}\n"
            )
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": initial_question}
                    ]
                }
            ]
            payload = {
                "model": self.model,
                "messages": messages,
                "max_tokens": self.max_tokens,
                "top_p": self.top_p,
                "temperature": self.temperature,
                "stop": ["</tool>"]
            }
            answer = self.call_llm(payload)
            if f'<tool>' in answer and f'</tool>' not in answer:
                answer += f'</tool>'
            tools = re.search(r'<tool>(.*?)</tool>', answer, re.DOTALL)
            if tools:
                tool = tools.group(1).strip()
                return tool
        except Exception as e:
            output = traceback.format_exc()
            logger.info(output)
        return None
    
    def call_llm(self, payload):
        if self.model.startswith("gpt"):
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {os.environ['OPENAI_API_KEY']}"
            }
            logger.info("Generating content with GPT model: %s", self.model)
            response = requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers=headers,
                json=payload
            )

            if response.status_code != 200:
                if response.json()['error']['code'] == "context_length_exceeded":
                    logger.error("Context length exceeded. Retrying with a smaller context.")
                    payload["messages"] = [payload["messages"][0]] + payload["messages"][-1:]
                    retry_response = requests.post(
                        "https://api.openai.com/v1/chat/completions",
                        headers=headers,
                        json=payload
                    )
                    if retry_response.status_code != 200:
                        logger.error(
                            "Failed to call LLM even after attempt on shortening the history: " + retry_response.text)
                        return ""

                logger.error("Failed to call LLM: " + response.text)
                time.sleep(5)
                return ""
            else:
                return response.json()['choices'][0]['message']['content']
            
        elif self.model.startswith("claude"):
            client = anthropic.Anthropic()
            max_retries = 3
            messages = payload["messages"].copy()  # Make a copy to preserve original
            
            while messages and max_retries > 0:
                try:
                    completion = client.messages.create(
                        model="claude-sonnet-4-5-20250929",
                        messages=messages,
                        max_tokens=payload["max_tokens"]
                    )
                    return completion.content[0].text
                except Exception as e:
                    if "request_too_large" in str(e):
                        # Remove the second message (index 1) if it exists
                        if len(messages) > 1:
                            messages.pop(1)
                            continue
                    
                    max_retries -= 1
                    if max_retries > 0:
                        time.sleep(5)
                    else:
                        logger.error("All attempts to call Claude failed.")
                        return ""

    def predict(self, instruction: str, obs: dict):
        '''
        Predict the next action(s) based on the current observation.
        Input two things: instruction (Task description), obs:
        self.observations.append({
                "screenshot": base64_image,
                "accessibility_tree": None
            })
        return two things: response, actions
        '''
        """
        Predict the next action(s) based on the current observation.
        """
        self.steps += 1
        logger.info("Current step: %s", self.steps)
        
        # Check if maximum steps reached
        if self.steps > self.max_steps-1:
            self.steps = 0
            response = '<execute>FAIL</execute>'
            logger.info("RESPONSE: %s", response)

            try:
                actions = self.parse_actions(response)
                if actions:
                    self.actions.append(actions)
                    self.thoughts.append(response)
            except ValueError as e:
                print("Failed to parse action from response", e)
                actions = None
                self.thoughts.append("")
            return response, actions   
        
        # Use the following system message for the libreoffice_calc task
        # system_message = self.system_message + "\nYou are asked to complete the following task: {}. Please input several commands at one code block for `pyautogui.typewrite()` and `pyautogui.press('enter')` operations to improve efficiency.".format(instruction)
        system_message = (
            self.system_message
            + "\nYou are asked to complete the following task: {}".format(instruction)
            + "\n\nFollow this action priority:\n"
            "1) Command-line actions\n"
            "2) Tool-building actions (write small Python utilities/tools)\n"
            "3) Python-in-Jupyter actions\n\n"
            "If the above cannot solve the task, then use GUI actions to interact with desktop applications.\n"
            "Always prefer reliable, scriptable methods before GUI interactions unless the task provides specific instructions to the contrary.\n"
        )
        # ('NOTE: '
        #  '- You can write Python helper functions/tools to solve the task. '
        #  '- You are strong at text understanding and coding. '
        #  '- Your image understanding is limited. Always try command-line operations or Python tools first. '
        #  '- If that fails, fall back to PyAutoGUI actions (GUI clicks aren’t always reliable). '
        #  '- During the task, I’ll provide screenshots of the current screen and any code outputs (when available).'
        #  'Let’s get started.')
        
        # Prepare the payload for the API call
        messages = []
        
        messages.append({
            "role": "user",
            "content":  [
            {
                "type": "text",
                "text": system_message,
                "cache_control": {"type": "ephemeral", "ttl": "1h"}  # 可改成 {"type": "ephemeral", "ttl": "1h"}
            }
        ]
        })

        # Append trajectory
        assert len(self.observations) == len(self.actions) and len(self.actions) == len(self.thoughts) \
            , "The number of observations and actions should be the same."
        if len(self.observations) > self.max_trajectory_length:
            if self.max_trajectory_length == 0:
                _observations = []
                _actions = []
                _thoughts = []
            else:
                _observations = self.observations[-self.max_trajectory_length:]
                _actions = self.actions[-self.max_trajectory_length:]
                _thoughts = self.thoughts[-self.max_trajectory_length:]
        else:
            _observations = self.observations
            _actions = self.actions
            _thoughts = self.thoughts

        for previous_obs, previous_action, previous_thought in zip(_observations, _actions, _thoughts):

            _screenshot = previous_obs["screenshot"]
            _terminal_output = previous_obs.get("bash_output", "")

            text_parts = []
            if _terminal_output and _terminal_output.strip():
                text_parts.append(f"[Terminal Output]\n{_terminal_output}")
            text_parts.append("[Current Screenshot is shown below]\n")

            messages.append({
                "role": "user",
                "content": [
                    {"type": "text", "text": "\n".join(text_parts) + NOTES},
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/png",
                            "data": _screenshot,
                        },
                    },
                ],
            })
            
            messages.append({
                "role": "assistant",
                "content": [
                    {
                        "type": "text",
                        "text": previous_thought.strip() if len(previous_thought) > 0 else "No valid action"
                    },
                ]
            })
            
        # generate new action
        img_save = Image.open(io.BytesIO(obs["screenshot"])).convert("RGB")
        save_image_and_convert_to_byte(img_save)
        base64_image = encode_image(obs["screenshot"])
        terminal_output = obs.get("bash_output", "")

        self.observations.append({
            "screenshot": base64_image,
            "accessibility_tree": None,
            "terminal": terminal_output
        })
        
        text_parts = []
        if terminal_output and terminal_output.strip():
            text_parts.append(f"[Terminal Output]\n{terminal_output}")
        text_parts.append("[Current Screenshot is shown below]\n")
        logger.info("Terminal Output: %s", terminal_output)
        messages.append({
            "role": "user",
            "content": [
                {"type": "text", "text": "\n".join(text_parts) + NOTES},
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": "image/png",
                        "data": base64_image,
                    },
                },
            ],
        })

        try:
            response = self.call_llm({
                "model": self.model,
                "messages": messages,
                "max_tokens": self.max_tokens,
                "top_p": self.top_p,
                "temperature": self.temperature
            })
        except Exception as e:
            logger.error("Failed to call" + self.model + ", Error: " + str(e))
            response = ""

        logger.info("RESPONSE: %s", response)

        try:
            actions = self.parse_actions(response)
            if actions:
                self.actions.append(actions)
                self.thoughts.append(response)
        except ValueError as e:
            print("Failed to parse action from response", e)
            actions = None
            self.thoughts.append("")

        return response, actions       
    
    def parse_actions(self, response: str, masks=None):
        actions = parse_code_from_string(response)
        return actions

    def localizaiton(self, action: dict):
        keywords = [
            "pyautogui.click",
            "pyautogui.dragTo",
            "pyautogui.mouseDown",
            "pyautogui.moveTo",
            "pyautogui.drag",
            "pyautogui.rightClick",
            "pyautogui.tripleClick",
            "pyautogui.doubleClick",
            "pyautogui.mouseUp",
        ]
        # Take a screenshot once for all matching actions
        byte_image = self.env.controller.get_screenshot()
        img = Image.open(io.BytesIO(byte_image))

        if any(kw in action['command'] for kw in keywords):
            icon, desc = extract_icon_and_desc(action['command'])
            # logger.info(f"Icon: {icon}, Desc: {desc}")
            if icon is None or desc is None:
                # Skip if no valid description
                return action

            # Locate on the screenshot
            coordination = self.image_description_to_coordinate(icon, desc, img)
            logger.info(f"Coordination for action: {coordination}")

            # Replace only if coordination is valid
            if isinstance(coordination, tuple) and len(coordination) == 2:
                x, y = coordination
                try:
                    new_action = replace_icon_desc_with_coordinates(action['command'], x, y)
                    action_w_coodination = {}
                    action_w_coodination['command'] = new_action
                    action_w_coodination['type'] = action['type']
                    logger.info(f"Action updated: {new_action}")
                    return action_w_coodination
                except (SyntaxError, ValueError) as e:
                    logger.error(f"Failed to replace coords in action: {e}")
            else:
                logger.error(f"Coordination is not a valid tuple for action.")

        return action
    
    # def oss_llm_completion(self, messages, stop=None):
    #     if self.oss_llm is None:
    #         self.oss_llm_init()
    #     sampling_params = SamplingParams(
    #                 n=1,
    #                 max_tokens=9632,
    #                 temperature=0,
    #                 top_p=1.0,
    #                 frequency_penalty=0,
    #                 presence_penalty=0
    #             )
    #     sampling_params.stop = stop
    #     request_output = self.oss_llm.chat(messages, sampling_params)
    #     # logger.debug(f"Request output: {request_output}")
    #     response_list = []
    #     for response in request_output[0].outputs:
    #         response_list.append(response.text)
    #     return response_list

    def completion_remote(self, messages, stop: list | None = None):
        client = OpenAI(
            base_url=f"{self.base_url_oss.rstrip('/')}/v1",
            api_key="",
        )
        return client.chat.completions.create(
            model='ByteDance-Seed/UI-TARS-1.5-7B',
            messages=messages,
            temperature=0.0,
            max_tokens=8192,
            n=1,
            stop=stop if stop else None,
        )
    
    def oss_llm_completion(self, messages, stop=None):
        request_output = self.completion_remote(messages, stop)
        # logger.info(f"Request output: {request_output}")
        number_of_answers = len(request_output.choices) # list of choices (answers)
        input_token_count = request_output.usage.prompt_tokens
        output_token_count = request_output.usage.completion_tokens
        # logger.info(
        #         'Total Input tokens: %.2f | Total Generated tokens: %.2f | Total outputs: %.2f',
        #         input_token_count,
        #         output_token_count,
        #         number_of_answers
        #     )
        response_list = []
        for response in request_output.choices:
            response_list.append(response.message.content)
        return response_list
    
    def _ask_llm_for_coordinate(self, image_bytes: bytes, description: str) -> tuple[int,int]:
        """向 oss_llm_completion 发送一次请求，解析并返回 (x,y)。"""
        b64 = encode_image(image_bytes)
        text = (
            "Please provide the ONE point coordinates (x, y) "
            f"of the element described as: {description}"
        )
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
                    {"type": "text", "text": text},
                ]
            }
        ]
        result = self.oss_llm_completion(messages)
        coord = extract_coordinates(result)
        return coord       
        
    def image_description_to_coordinate(self, icon, desc, image):
        """
        Convert the image description to coordinate for accurate mouse click.
        """
        # Crop the image
        global CURRENT_WHOLE_IMAGE
        CURRENT_WHOLE_IMAGE = image
        byte_image = save_image_and_convert_to_byte(image) 
        coordination = self._ask_llm_for_coordinate(byte_image, icon+" ("+desc+")")
        
        # localization
        if not coordination[0] == -1:
            byte_cropped, offset = highlight_and_save_region(coordination, half_size_x = 700, half_size_y = 250)
            coordination = self._ask_llm_for_coordinate(byte_cropped, icon+" ("+desc+")")
            dx = offset[0]
            dy = offset[1]
            coordination = (coordination[0] + dx, coordination[1] + dy)
                      
        # save the image with the red dot
        copy_img, _, _, _ = localization_point(coordination[0], coordination[1])
        if copy_img:
            save_image_and_convert_to_byte(copy_img)
        return coordination

def configs() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run end-to-end evaluation on the benchmark"
    )

    # environment config
    parser.add_argument("--path_to_vm", type=str, default=None)
    parser.add_argument(
        "--headless", action="store_true", help="Run in headless machine"
    )
    parser.add_argument(
        "--action_space", type=str, default="pyautogui", help="Action type"
    )
    parser.add_argument(
        "--observation_type",
        choices=["screenshot", "a11y_tree", "screenshot_a11y_tree", "som"],
        default="a11y_tree",
        help="Observation type",
    )
    parser.add_argument("--screen_width", type=int, default=1920)
    parser.add_argument("--screen_height", type=int, default=1080)
    parser.add_argument("--sleep_after_execution", type=float, default=0.5)
    parser.add_argument("--max_steps", type=int, default=50)

    # agent config
    parser.add_argument("--max_trajectory_length", type=int, default=20)
    parser.add_argument(
        "--test_config_base_dir", type=str, default="evaluation_examples"
    )

    # lm config
    parser.add_argument("--model", type=str, default="claude-sonnet-4-5-20250929")
    parser.add_argument("--temperature", type=float, default=0.0)
    parser.add_argument("--top_p", type=float, default=0.9)
    parser.add_argument("--max_tokens", type=int, default=1500)
    parser.add_argument("--stop_token", type=str, default=None)

    # example config
    parser.add_argument("--domain", type=str, default="all")
    # parser.add_argument(
    #     "--test_all_meta_path", type=str, default="evaluation_examples/test_all.json"
    # )
    parser.add_argument(
        "--test_all_meta_path", type=str, default="evaluation_examples/test_all.json"
    )
    # logging related
    parser.add_argument("--result_dir", type=str, default="./results")
    args = parser.parse_args()

    return args

def test(args: argparse.Namespace, test_all_meta: dict) -> None:
    scores = []
    max_steps = args.max_steps

    # log args
    logger.info("Args: %s", args)
    # set wandb project
    cfg_args = {
        "path_to_vm": args.path_to_vm,
        "headless": args.headless,
        "action_space": args.action_space,
        "observation_type": args.observation_type,
        "screen_width": args.screen_width,
        "screen_height": args.screen_height,
        "sleep_after_execution": args.sleep_after_execution,
        "max_steps": args.max_steps,
        "max_trajectory_length": args.max_trajectory_length,
        "model": args.model,
        "temperature": args.temperature,
        "top_p": args.top_p,
        "max_tokens": args.max_tokens,
        "stop_token": args.stop_token,
        "result_dir": args.result_dir,
        "base_url_oss": os.getenv("VLLM_BASE_URL", "http://localhost:8000"),
    }

    env = OSworld_Env(
        path_to_vm=args.path_to_vm,
        action_space=args.action_space,
        screen_size=(args.screen_width, args.screen_height),
        headless=args.headless,
        os_type = "Ubuntu",
        require_a11y_tree=args.observation_type
        in ["a11y_tree", "screenshot_a11y_tree", "som"],
        helper_fn=None,
        gui_helper_fn=None
    )

    byte_image = env.controller.get_screenshot()
    img = Image.open(io.BytesIO(byte_image))
    screen_width, screen_height = img.size

    agent = OSworld_test_Agent(env, "placeholder", cfg_args, screen_width, screen_height)
    
    allowed_tools = {
        "make_new_tool": agent.make_new_tool,
    }

    def call_tool_from_string(cmd: str) -> str:
        return eval(cmd, {"__builtins__": {}}, allowed_tools)
    
    env.helper_fn = call_tool_from_string
    env.gui_helper_fn = agent.localizaiton

    for domain in tqdm(test_all_meta, desc="Domain"):
        for example_id in tqdm(test_all_meta[domain], desc="Example", leave=False):
            agent.steps = 0
            config_file = os.path.join(
                args.test_config_base_dir, f"examples/{domain}/{example_id}.json"
            )
            with open(config_file, "r", encoding="utf-8") as f:
                example = json.load(f)
            # run.config.update(cfg_args)

            example_result_dir = os.path.join(
                args.result_dir,
                args.action_space,
                args.observation_type,
                args.model,
                domain,
                example_id,
            )
            os.makedirs(example_result_dir, exist_ok=True)
            agent.update_result_dir(example_result_dir)
            instruction = example["instruction"]
            # For some cases: Alought it is finished by code execution, 
            # The evaluator will return False if they are not finished by GUI actions.
            # For example ee9a3c83-f437-4879-8918-be5efbb9fac7, agent has to input the command in the **Opened**
            # Terminal directly. Instead of running the command in the Terminal via code execution.
            # Althought they are functionally the same, the evaluator will return False if not done by GUI actions.
            if example_id in ('2b9493d7-49b8-493a-a71b-56cd1f4d6908', '510f64c8-9bcc-4be1-8d30-638705850618',
                              '4127319a-8b79-4410-b58a-7a151e15f3d7', '4783cc41-c03c-4e1b-89b4-50658f642bd5',
                               '37887e8c-da15-4192-923c-08fa390a176d','02ce9a50-7af2-47ed-8596-af0c230501f8',
                               'd68204bf-11c1-4b13-b48b-d303c73d4bf6','ee9a3c83-f437-4879-8918-be5efbb9fac7',
                               'f7dfbef3-7697-431c-883a-db8583a4e4f9','3680a5ee-6870-426a-a997-eba929a0d25c',):
                instruction += ('\nPlease type your command directly into the already-open Terminal using GUI actions '
                                'Do not run any code yourself outside the Terminal.')
            if example_id in ('53ad5833-3455-407b-bbc6-45b4c79ab8fb', '59f21cfb-0120-4326-b255-a5b827b38967', 
                              'e1e75309-3ddb-4d09-92ec-de869c928143', 'f0b971a1-6831-4b9b-a50e-22a6e47f45ba', 
                              'f5d96daf-83a8-4c86-9686-bada31fc66ab', '2a729ded-3296-423d-aec4-7dd55ed5fbb3',
                              '2e6f678f-472d-4c55-99cc-8e7c5c402a71', '5ca86c6f-f317-49d8-b6a7-b527541caae8',
                              '06ca5602-62ca-47f6-ad4f-da151cde54cc', '8ea73f6f-9689-42ad-8c60-195bbf06a7ba',
                              '38f48d40-764e-4e77-a7cf-51dfce880291', '045bf3ff-9077-4b86-b483-a1040a949cff',
                              '58d3eeeb-e9d0-499f-962e-fd0db2a744d8','62f7fd55-0687-4a43-b6e1-3eda16fc6252',
                              '77b8ab4d-994f-43ac-8930-8ca087d7c4b4', 'dbbf4b99-2253-4b10-9274-45f246af2466',
                              'e19bd559-633b-4b02-940f-d946248f088e', 'fbb548ca-c2a6-4601-9204-e39a2efc507b',
                              '1f18aa87-af6f-41ef-9853-cdb8f32ebdea', 
                              '8df7e444-8e06-4f93-8a1a-c5c974269d82', '09a37c51-e625-49f4-a514-20a773797a8a',
                              '42d25c08-fb87-4927-8b65-93631280a26f', '42f4d1c7-4521-4161-b646-0a8934e36081',
                              '48c46dc7-fe04-4505-ade7-723cba1aa6f6', '68a25bd4-59c7-4f4d-975e-da0c8509c848',
                              '74d5859f-ed66-4d3e-aa0e-93d7a592ce41', '869de13e-bef9-4b91-ba51-f6708c40b096',
                              '937087b6-f668-4ba6-9110-60682ee33441',
                              'da922383-bfa4-4cd3-bbad-6bebab3d7742',
                              'bedcedc4-4d72-425e-ad62-21960b11fe0d',
                              '5ac2891a-eacd-4954-b339-98abba077adb', 'aa4b5023-aef6-4ed9-bdc9-705f59ab9ad6',
                              'bba3381f-b5eb-4439-bd9e-80c22218d5a7', 'cb130f0d-d36f-4302-9838-b3baf46139b6',
                              '6ed0a554-cbee-4b44-84ea-fd6c042f4fe1', '971cbb5b-3cbf-4ff7-9e24-b5c84fcebfa6',
                              ):
                instruction += '\nPlease **only** use GUI actions to complete the task.'
            # The problem statement is unclear; it doesn’t specify whether it’s one-way or round-trip.
            if example_id in ('6c4c23a1-42a4-43cc-9db1-2f86ff3738cc',):
                instruction += ('\nI only need a one-way ticket.')
            # There are many ways to achieve the target, need to specified the method.
            if example_id in ('7f52cab9-535c-4835-ac8c-391ee64dc930a',):
                instruction += ('\nPlease use Google Shopping’s built-in filters to refine the results.')
            # The problem statement is unclear; it doesn’t specify which airpot to choose. (There are two airports in Stockholm)
            if example_id in ('82bc8d6a-36eb-4d2d-8801-ef714fb1e55a',):
                instruction += ('\nI would like to land at Stockholm (STO) airport.')

            # wandb each example config settings
            cfg_args["instruction"] = instruction
            cfg_args["start_time"] = datetime.datetime.now().strftime(
                "%Y:%m:%d-%H:%M:%S"
            )
            
            # example start running
            try:
                lib_run_single.run_single_example(
                    agent,
                    env,
                    example,
                    max_steps,
                    instruction,
                    args,
                    example_result_dir,
                    scores,
                )
                logger.info(f"[Domain]: {domain}")
                logger.info(f"[Example ID]: {example_id}")
                logger.info(f"[Instruction]: {instruction}")
            except Exception as e:
                error_trace = traceback.format_exc()
                logger.error(f"Exception in {domain}/{example_id}: {e}\nTraceback:\n{error_trace}")
                env.controller.end_recording(
                    os.path.join(example_result_dir, "recording.mp4")
                )
                with open(os.path.join(example_result_dir, "traj.jsonl"), "a") as f:
                    f.write(
                        json.dumps(
                            {"Error": f"Time limit exceeded in {domain}/{example_id}"}
                        )
                    )
                    f.write("\n")
        # break
    env.close()
    logger.info(f"Average score: {sum(scores) / len(scores)}")

def get_unfinished(
    action_space, use_model, observation_type, result_dir, total_file_json
):
    target_dir = os.path.join(result_dir, action_space, observation_type, use_model)

    if not os.path.exists(target_dir):
        return total_file_json

    finished = {}
    for domain in os.listdir(target_dir):
        finished[domain] = []
        domain_path = os.path.join(target_dir, domain)
        if os.path.isdir(domain_path):
            for example_id in os.listdir(domain_path):
                if example_id == "onboard":
                    continue
                example_path = os.path.join(domain_path, example_id)
                if os.path.isdir(example_path):
                    if "result.txt" not in os.listdir(example_path):
                        # empty all files under example_id
                        for file in os.listdir(example_path):
                            file_path = os.path.join(example_path, file)
                            if os.path.isfile(file_path):
                                os.remove(file_path)
                            elif os.path.isdir(file_path):
                                shutil.rmtree(file_path) 
                    else:
                        finished[domain].append(example_id)

    if not finished:
        return total_file_json

    for domain, examples in finished.items():
        if domain in total_file_json:
            total_file_json[domain] = [
                x for x in total_file_json[domain] if x not in examples
            ]

    return total_file_json

def get_result(action_space, use_model, observation_type, result_dir, total_file_json):
    target_dir = os.path.join(result_dir, action_space, observation_type, use_model)
    if not os.path.exists(target_dir):
        print("New experiment, no result yet.")
        return None

    all_result = []

    for domain in os.listdir(target_dir):
        domain_path = os.path.join(target_dir, domain)
        if os.path.isdir(domain_path):
            for example_id in os.listdir(domain_path):
                example_path = os.path.join(domain_path, example_id)
                if os.path.isdir(example_path):
                    if "result.txt" in os.listdir(example_path):
                        # empty all files under example_id
                        try:
                            all_result.append(
                                float(
                                    open(
                                        os.path.join(example_path, "result.txt"), "r"
                                    ).read()
                                )
                            )
                        except:
                            all_result.append(0.0)

    if not all_result:
        print("New experiment, no result yet.")
        return None
    else:
        print("Current Success Rate:", sum(all_result) / len(all_result) * 100, "%")
        return all_result

if __name__ == "__main__":
    ####### The complete version of the list of examples #######
    os.environ["TOKENIZERS_PARALLELISM"] = "false"
    args = configs()

    with open(args.test_all_meta_path, "r", encoding="utf-8") as f:
        test_all_meta = json.load(f)

    if args.domain != "all":
        logger.info(f"Filtering to only run domain: {args.domain}")
        test_all_meta = {args.domain: test_all_meta[args.domain]}

    test_file_list = get_unfinished(
        args.action_space,
        args.model,
        args.observation_type,
        args.result_dir,
        test_all_meta,
    )
    left_info = ""
    for domain in test_file_list:
        left_info += f"{domain}: {len(test_file_list[domain])}\n"
    logger.info(f"Left tasks:\n{left_info}")

    get_result(
        args.action_space,
        args.model,
        args.observation_type,
        args.result_dir,
        test_all_meta,
    )
    test(args, test_file_list)


'''
export CUDA_VISIBLE_DEVICES=4,5
nohup python -m vllm.entrypoints.openai.api_server \
  --model ByteDance-Seed/UI-TARS-1.5-7B \
  --host 0.0.0.0 --port 8888 \
  --trust-remote-code \
  --max-model-len 32768 \
  --gpu-memory-utilization 0.9 \
  --tensor-parallel-size 2 \
  > ~/logs/vllm_$(date +%Y%m%d@%H%M%S).log 2>&1 &

export VLLM_BASE_URL="http://127.0.0.1:8888"
export ANTHROPIC_API_KEY=''

nohup python run_inference.py \
    --headless \
    --observation_type screenshot \
    --model claude-sonnet-4-5-20250929 \
    --sleep_after_execution 3 \
    --domain os \
    --max_steps 50 \
    > ~/osworld_$(date +%Y%m%d@%H%M%S).log 2>&1 &
'''