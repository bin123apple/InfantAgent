from __future__ import annotations

import os
import re
import asyncio
import traceback
from infant.config import config
from infant.llm.llm_api_base import LLM_API_BASED
from infant.agent.memory.memory import IPythonRun
from infant.util.logger import infant_logger as logger
from infant.util.run_code_locally import run_code_in_subprocess
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from infant.agent.agent import Agent

# TODO: We need to write a unittest for this
def make_new_tool(functionality: str, function_name: str,
                        function_inputs: list, function_outputs: str) -> str:
    '''
    Make a new tool based on the request from the Agent
    If the new tool is created:
        return the created tool
    if not:
        return None
    '''
    try:
        tm_parameter = config.get_litellm_params(overrides = config.tm_llm)
        tm_llm = LLM_API_BASED(tm_parameter)
        if tm_llm.api_key is None:
            tm_llm.api_key = os.getenv("ANTHROPIC_API_KEY")
            print(f'tm_llm.api_key:{tm_llm.api_key}')
        
        example = '''USER:
I need a function that can create a new sheet in an existing Excel file.

Functionality:  
- Open an existing Excel file.  
- Create a new sheet with the specified name.  
- Save the updated Excel file.  
- If the sheet already exists, raise a ValueError.  

Function name: create_excel_page  
Input: file_name (str), page_name (str)  
Output: None  

ASSISTANT:
Here is the complete function wrapped inside the <tool>...</tool> tag:

<tool>
import openpyxl

def create_excel_page(file_name: str, page_name: str) -> None:
    """
    Create a new sheet in the given Excel file.

    Args:
        file_name (str): Path to the Excel file.
        page_name (str): Name of the new sheet to be created.

    Returns:
        None

    Raises:
        ValueError: If the sheet already exists.
    """
    wb = openpyxl.load_workbook(file_name)
    if page_name in wb.sheetnames:
        raise ValueError(f"Sheet '{page_name}' already exists.")
    wb.create_sheet(title=page_name)
    wb.save(file_name)
</tool>'''

        initial_question = (
                            f"Please help me to write a new python function based on the following instruction: "
                            f"The function should provide the following functionality. \n{functionality}\n"
                            f"The name of the function should be: \n{function_name}\n" 
                            f"The input of the function should be: \n{function_inputs}\n"
                            f"The output of the function should be: \n{function_outputs}\n"
                            "Please verify your function thoroughly. "
                            "Finally, please only provide the complete function and wrap it inside the <tool>...</tool> tag."
                            "Do not include anything else inside the <tool>...</tool> tag (such as unit tests or usage examples). "
                            "Only place the function itself within the tag. "
                            f"Here is an example: \n{example}\n"
        )
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": initial_question}
                ]
            }
        ]
        answer, _ = tm_llm.completion(messages=messages,stop=['</tool>'])
        if f'<tool>' in answer and f'</tool>' not in answer:
            answer += f'</tool>'
        tools = re.search(r'<tool>(.*?)</tool>', answer, re.DOTALL)
        if tools:
            tool = tools.group(1).strip()
            return tool
    except Exception as e:
        output = traceback.format_exc()
        logger.info(output)
    return None

async def make_tool(agent: Agent, memory: IPythonRun) -> str:
    ''' 
    Unify the interface, execute this function will run the memory.action 
    1. add the new tool to the toolsets. 
    2. generate the memory.result
        if the tool is created successfully, update the memory.result with the new tool information.
        if the tool creation fails, update the memory.result with an error message.
    '''
    computer = agent.computer
    code = memory.code
    prefix = 'from infant.helper_functions.toolmaker_helper_function import make_new_tool'
    rc, stdout, stderr = await run_code_in_subprocess(f'{prefix}\nprint({code})',
                                                      extra_env={"ANTHROPIC_API_KEY": agent.tm_llm.api_key})
    print(f'stdout: {stdout.decode()}') # actual return
    print(f'stderr: {stderr.decode()}') # error message or log
    tool = stdout.decode().strip() if rc == 0 else None
    if tool: # the tool is created successfully
        new_tool = (
            f'The new tool is created, here is its detailed implementation:\n{tool}\n'
            'If you encounter any issues during use, please let me know.'
        )
        computer.run_python(tool) # add the new tool to the toolsets
        memory.result = new_tool
    else: # the tool creation fails
        tool_creation_error = stderr.decode().strip() if stderr else 'Unknown error'
        error_msg = (
            f'Sorry, I was unable to create the tool due to the following error: {tool_creation_error}\n'
            'Please review the functionality requirements and try again.'
        )
        memory.result = error_msg
    return memory
        